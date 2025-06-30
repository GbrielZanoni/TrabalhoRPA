import tkinter as tk
from tkinter import ttk
from tkinter import messagebox # Movido para o topo
from tkinter import filedialog # Para selecionar arquivos
from datetime import datetime
import pandas as pd
import shutil # Para copiar arquivos
from openpyxl import load_workbook # Embora o pandas possa lidar com a escrita/anexação sem carregar explicitamente com openpyxl para .xlsx
from fpdf import FPDF 
# import smtplib # Comentado - E-mail removido
import os # Ainda usado para manipulação de caminhos de arquivos (fotos)
# from email.mime.multipart import MIMEMultipart # Comentado - E-mail removido
# from email.mime.text import MIMEText # Comentado - E-mail removido
# from email.mime.base import MIMEBase # Comentado - E-mail removido
# from email import encoders # Comentado - E-mail removido

class ChecklistApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Checklist Digital - Energral")
        self.root.geometry("800x700") # Ajustado para melhor visualização inicial

        # Estilo para os frames
        style = ttk.Style()
        style.configure("TFrame", padding=10, relief="flat")
        style.configure("TLabel", padding=5)
        style.configure("TButton", padding=5)
        style.configure("TRadiobutton", padding=5)
        style.configure("TCheckbutton", padding=5)

        # --- Configuração da Barra de Rolagem Principal ---
        # Criar um Canvas e uma Scrollbar
        canvas = tk.Canvas(root)
        scrollbar = ttk.Scrollbar(root, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # Frame rolável que conterá todo o conteúdo do checklist
        # Este é o frame que antes era o 'main_frame' direto no 'root'
        self.scrollable_frame = ttk.Frame(canvas, style="TFrame")
        
        # Adicionar o frame rolável ao canvas
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

        # Atualizar a região de rolagem do canvas quando o tamanho do scrollable_frame mudar
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )
        # --- Fim da Configuração da Barra de Rolagem ---


        # Container principal (agora é o scrollable_frame)
        # main_frame = ttk.Frame(root, style="TFrame") # Linha antiga
        # main_frame.pack(fill=tk.BOTH, expand=True) # Linha antiga
        # Referenciar main_frame ao scrollable_frame para manter o resto do código
        main_frame = self.scrollable_frame


        # Título do Aplicativo
        app_title = ttk.Label(main_frame, text="Checklist Digital para Técnicos da Energral", font=("Arial", 16, "bold"))
        app_title.pack(pady=(10, 20), padx=10) # Adicionado padx para não colar na borda com scrollbar

        # --- Seções do Checklist ---
        # Cada seção será um ttk.Labelframe para melhor organização visual

        # 1. Dados da Visita
        dados_visita_frame = ttk.Labelframe(main_frame, text="1. Dados da Visita", padding=10)
        dados_visita_frame.pack(fill=tk.X, padx=10, pady=5)
        self.create_dados_visita_section(dados_visita_frame)

        # 2. Estado dos Equipamentos
        estado_equipamentos_frame = ttk.Labelframe(main_frame, text="2. Estado dos Equipamentos", padding=10)
        estado_equipamentos_frame.pack(fill=tk.X, padx=10, pady=5)
        self.create_estado_equipamentos_section(estado_equipamentos_frame)

        # 3. Segurança
        seguranca_frame = ttk.Labelframe(main_frame, text="3. Segurança", padding=10)
        seguranca_frame.pack(fill=tk.X, padx=10, pady=5)
        self.create_seguranca_section(seguranca_frame)

        # 4. Incidentes
        incidentes_frame = ttk.Labelframe(main_frame, text="4. Incidentes (se houver falha)", padding=10)
        incidentes_frame.pack(fill=tk.X, padx=10, pady=5)
        self.create_incidentes_section(incidentes_frame)

        # 5. Observações Livres
        observacoes_frame = ttk.Labelframe(main_frame, text="5. Observações Livres", padding=10)
        observacoes_frame.pack(fill=tk.X, padx=10, pady=5) # fill=tk.BOTH, expand=True se quiser que expanda verticalmente também
        self.create_observacoes_section(observacoes_frame)

        # Botão de Salvar/Submeter
        submit_button = ttk.Button(main_frame, text="Salvar Checklist", command=self.submit_checklist)
        submit_button.pack(pady=20, padx=10)

    def create_dados_visita_section(self, parent_frame):
        # Frame interno para melhor layout dos widgets
        frame = ttk.Frame(parent_frame)
        frame.pack(fill=tk.X)

        # Data e Hora (Automático)
        ttk.Label(frame, text="Data e Hora:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        self.data_hora_var = tk.StringVar(value=datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        ttk.Entry(frame, textvariable=self.data_hora_var, state="readonly", width=30).grid(row=0, column=1, sticky=tk.W, padx=5, pady=2)

        # Local da Subestação (Lista)
        ttk.Label(frame, text="Local da Subestação:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        subestacoes = ["Subestação Centro", "Subestação Norte", "Subestação Sul", "Subestação Leste", "Subestação Oeste"]
        self.local_subestacao_var = tk.StringVar()
        local_combobox = ttk.Combobox(frame, textvariable=self.local_subestacao_var, values=subestacoes, width=28)
        local_combobox.grid(row=1, column=1, sticky=tk.W, padx=5, pady=2)
        local_combobox.current(0) # Define um valor padrão

        # Nome do Técnico (Editável)
        ttk.Label(frame, text="Nome do Técnico:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
        self.nome_tecnico_var = tk.StringVar() # Sem valor padrão, ou pode ser um placeholder
        nome_tecnico_entry = ttk.Entry(frame, textvariable=self.nome_tecnico_var, width=30)
        nome_tecnico_entry.grid(row=2, column=1, sticky=tk.W, padx=5, pady=2)
        # Opcional: adicionar um placeholder leve se o ttk não suportar diretamente
        # nome_tecnico_entry.insert(0, "Digite o nome do técnico") 
        # nome_tecnico_entry.bind("<FocusIn>", lambda args: nome_tecnico_entry.delete('0', 'end') if nome_tecnico_entry.get() == "Digite o nome do técnico" else None)
        # nome_tecnico_entry.bind("<FocusOut>", lambda args: nome_tecnico_entry.insert(0, "Digite o nome do técnico") if nome_tecnico_entry.get() == "" else None)


        # Tipo de Inspeção (Rotina ou Emergência)
        ttk.Label(frame, text="Tipo de Inspeção:").grid(row=3, column=0, sticky=tk.W, padx=5, pady=2)
        self.tipo_inspecao_var = tk.StringVar(value="Rotina")
        ttk.Radiobutton(frame, text="Rotina", variable=self.tipo_inspecao_var, value="Rotina").grid(row=3, column=1, sticky=tk.W, padx=5)
        ttk.Radiobutton(frame, text="Emergência", variable=self.tipo_inspecao_var, value="Emergência").grid(row=3, column=2, sticky=tk.W, padx=5)

    def create_estado_equipamentos_section(self, parent_frame):
        # Placeholder
        # Helper para criar par de Checkbuttons OK/Falha
        def create_ok_falha_pair(parent, label_text, row):
            var = tk.StringVar(value="OK") # Valor padrão OK
            ttk.Label(parent, text=label_text).grid(row=row, column=0, sticky=tk.W, padx=5, pady=2)
            
            ok_button = ttk.Radiobutton(parent, text="✔️ OK", variable=var, value="OK")
            ok_button.grid(row=row, column=1, sticky=tk.W, padx=5)
            
            falha_button = ttk.Radiobutton(parent, text="❌ Falha", variable=var, value="Falha")
            falha_button.grid(row=row, column=2, sticky=tk.W, padx=5)
            return var

        # Dicionário para armazenar as variáveis de estado dos equipamentos
        self.estado_equipamentos_vars = {}

        # Transformador
        transformador_frame = ttk.LabelFrame(parent_frame, text="Transformador", padding=5)
        transformador_frame.pack(fill=tk.X, padx=5, pady=5)
        self.estado_equipamentos_vars["oleo_transformador"] = create_ok_falha_pair(transformador_frame, "Nível de óleo dentro do padrão:", 0)
        self.estado_equipamentos_vars["vazamentos_ruidos_transformador"] = create_ok_falha_pair(transformador_frame, "Sem vazamentos ou ruídos anormais:", 1)
        self.estado_equipamentos_vars["temperatura_transformador"] = create_ok_falha_pair(transformador_frame, "Temperatura dentro da faixa operacional:", 2)

        # Disjuntores
        disjuntores_frame = ttk.LabelFrame(parent_frame, text="Disjuntores", padding=5)
        disjuntores_frame.pack(fill=tk.X, padx=5, pady=5)
        self.estado_equipamentos_vars["arco_superaquecimento_disjuntor"] = create_ok_falha_pair(disjuntores_frame, "Sem sinais de arco elétrico ou superaquecimento:", 0)
        self.estado_equipamentos_vars["operacao_manual_disjuntor"] = create_ok_falha_pair(disjuntores_frame, "Operação manual livre (sem travamento):", 1)

        # Barramentos e Conexões
        barramentos_frame = ttk.LabelFrame(parent_frame, text="Barramentos e Conexões", padding=5)
        barramentos_frame.pack(fill=tk.X, padx=5, pady=5)
        self.estado_equipamentos_vars["corrosao_folgas_barramento"] = create_ok_falha_pair(barramentos_frame, "Sem corrosão ou folgas:", 0)
        self.estado_equipamentos_vars["isoladores_integros_barramento"] = create_ok_falha_pair(barramentos_frame, "Isoladores íntegros (sem trincas):", 1)

    def create_seguranca_section(self, parent_frame):
        # Placeholder
        # Dicionário para armazenar as variáveis de segurança
        self.seguranca_vars = {}

        # Frame interno para melhor layout
        frame = ttk.Frame(parent_frame)
        frame.pack(fill=tk.X)

        # EPIs Verificados
        epis_frame = ttk.LabelFrame(frame, text="EPIs Verificados", padding=5)
        epis_frame.pack(fill=tk.X, padx=5, pady=5)

        self.seguranca_vars["luva_isolante"] = tk.BooleanVar(value=True) # Padrão marcado
        ttk.Checkbutton(epis_frame, text="Luva isolante ✔️", variable=self.seguranca_vars["luva_isolante"]).pack(anchor=tk.W)

        self.seguranca_vars["botina_seguranca"] = tk.BooleanVar(value=True) # Padrão marcado
        ttk.Checkbutton(epis_frame, text="Botina de segurança ✔️", variable=self.seguranca_vars["botina_seguranca"]).pack(anchor=tk.W)

        self.seguranca_vars["capacete"] = tk.BooleanVar(value=True) # Padrão marcado
        ttk.Checkbutton(epis_frame, text="Capacete ✔️", variable=self.seguranca_vars["capacete"]).pack(anchor=tk.W)

        # Sinalização
        sinalizacao_frame = ttk.LabelFrame(frame, text="Sinalização", padding=5)
        sinalizacao_frame.pack(fill=tk.X, padx=5, pady=5)

        self.seguranca_vars["placas_alerta"] = tk.BooleanVar(value=True) # Padrão marcado
        ttk.Checkbutton(sinalizacao_frame, text="Placas de alerta visíveis e legíveis", variable=self.seguranca_vars["placas_alerta"]).pack(anchor=tk.W)

        self.seguranca_vars["area_livre"] = tk.BooleanVar(value=True) # Padrão marcado
        ttk.Checkbutton(sinalizacao_frame, text="Área livre de obstruções (matos, animais)", variable=self.seguranca_vars["area_livre"]).pack(anchor=tk.W)


    def create_incidentes_section(self, parent_frame):
        # Placeholder
        # Frame interno para melhor layout
        frame = ttk.Frame(parent_frame)
        frame.pack(fill=tk.X)

        # Dicionário para armazenar as variáveis de incidentes
        self.incidentes_vars = {}

        # Descrição do Problema
        ttk.Label(frame, text="Descrição do Problema:").grid(row=0, column=0, sticky=tk.NW, padx=5, pady=2)
        self.incidentes_vars["descricao"] = tk.Text(frame, height=4, width=60, wrap=tk.WORD)
        self.incidentes_vars["descricao"].grid(row=0, column=1, columnspan=2, sticky=tk.W, padx=5, pady=2)
        # Adicionando uma scrollbar para o campo de texto, caso o texto seja longo
        desc_scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.incidentes_vars["descricao"].yview)
        desc_scrollbar.grid(row=0, column=3, sticky=tk.NS, pady=2)
        self.incidentes_vars["descricao"]['yscrollcommand'] = desc_scrollbar.set


        # Botão para Foto (Real)
        self.incidentes_vars["foto_path"] = tk.StringVar() # Armazenará o caminho da foto copiada
        ttk.Button(frame, text="Selecionar Foto", command=self.selecionar_foto_incidente).grid(row=1, column=1, sticky=tk.W, padx=5, pady=5)
        self.foto_label = ttk.Label(frame, text="Nenhuma foto selecionada.", width=40, anchor="w") # Para mostrar o nome do arquivo
        self.foto_label.grid(row=1, column=2, sticky=tk.W, padx=5, pady=2)

        # Gravidade
        ttk.Label(frame, text="Gravidade:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
        self.incidentes_vars["gravidade"] = tk.StringVar(value="Baixa") # Padrão
        ttk.Radiobutton(frame, text="Baixa (pode aguardar manutenção programada)", variable=self.incidentes_vars["gravidade"], value="Baixa").grid(row=2, column=1, sticky=tk.W, padx=5)
        ttk.Radiobutton(frame, text="Alta (parada imediata necessária)", variable=self.incidentes_vars["gravidade"], value="Alta").grid(row=2, column=2, sticky=tk.W, padx=5)

        # Ação Tomada
        ttk.Label(frame, text="Ação Tomada:").grid(row=3, column=0, sticky=tk.W, padx=5, pady=2)
        acoes = [
            "Nenhuma ação imediata",
            "Isolamento do circuito",
            "Substituição de peça X",
            "Necessita análise especializada",
            "Outra (descrever em observações)"
        ]
        self.incidentes_vars["acao_tomada"] = tk.StringVar()
        acao_combobox = ttk.Combobox(frame, textvariable=self.incidentes_vars["acao_tomada"], values=acoes, width=40)
        acao_combobox.grid(row=3, column=1, columnspan=2, sticky=tk.W, padx=5, pady=2)
        acao_combobox.current(0) # Padrão

    def selecionar_foto_incidente(self):
        """Permite ao usuário selecionar um arquivo de imagem e o copia para um diretório local."""
        # Tipos de arquivo permitidos
        filetypes = (
            ("Imagens JPEG", "*.jpg *.jpeg"),
            ("Imagens PNG", "*.png"),
            ("Todos os arquivos", "*.*")
        )
        
        filepath = filedialog.askopenfilename(
            title="Selecionar foto do incidente",
            filetypes=filetypes
        )

        if not filepath: # Usuário cancelou a seleção
            return

        # Diretório para armazenar anexos
        anexos_dir = "anexos_checklist"
        if not os.path.exists(anexos_dir):
            try:
                os.makedirs(anexos_dir)
            except Exception as e:
                messagebox.showerror("Erro ao Criar Diretório", f"Não foi possível criar o diretório de anexos '{anexos_dir}':\n{e}")
                return
        
        # Cria um nome de arquivo único para evitar sobrescrita (opcional, mas bom)
        # Poderia ser mais robusto, com timestamp + nome original
        filename = os.path.basename(filepath)
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        destination_filename = f"{timestamp}_{filename}"
        destination_path = os.path.join(anexos_dir, destination_filename)

        try:
            shutil.copy(filepath, destination_path)
            self.incidentes_vars["foto_path"].set(destination_path) # Armazena o caminho do arquivo copiado
            self.foto_label.config(text=filename) # Mostra apenas o nome do arquivo original no label
            messagebox.showinfo("Foto Anexada", f"Foto '{filename}' anexada com sucesso!\nSalva como: {destination_path}")
            print(f"Foto selecionada: {filepath}")
            print(f"Foto copiada para: {destination_path}")
        except Exception as e:
            self.incidentes_vars["foto_path"].set("") # Limpa se a cópia falhar
            self.foto_label.config(text="Falha ao anexar foto.")
            messagebox.showerror("Erro ao Anexar Foto", f"Não foi possível copiar a foto para o diretório de anexos:\n{e}")
            print(f"Erro ao copiar foto: {e}")


    def create_observacoes_section(self, parent_frame):
        # Placeholder
        # Frame interno para melhor layout
        frame = ttk.Frame(parent_frame)
        frame.pack(fill=tk.BOTH, expand=True) # Expandir para preencher o espaço

        ttk.Label(frame, text="Comentários Adicionais:").pack(anchor=tk.W, padx=5, pady=(0,2))

        self.observacoes_text = tk.Text(frame, height=5, width=70, wrap=tk.WORD)
        self.observacoes_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=2)

        # Adicionando uma scrollbar para o campo de texto
        obs_scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.observacoes_text.yview)
        obs_scrollbar.pack(side=tk.RIGHT, fill=tk.Y) # Empacotar à direita
        self.observacoes_text['yscrollcommand'] = obs_scrollbar.set


    def submit_checklist(self):
        # A lógica de impressão direta foi removida e substituída pela coleta estruturada.
        collected_data = self._collect_data()

        # --- Impressão no console do dicionário estruturado (para depuração/log) ---
        print("--- DADOS DO CHECKLIST COLETADOS ---")
        import json # Importar json aqui para esta função específica
        print(json.dumps(collected_data, indent=2, ensure_ascii=False)) # Imprime o dicionário formatado
        print("\n--- FIM DOS DADOS COLETADOS ---")
        
        # Futuras chamadas para exportar para Excel, gerar PDF, enviar email virão aqui
        # ex: self.export_to_excel(collected_data)
        # ex: self.generate_pdf(collected_data)
        # ex: self.send_email_with_attachments(collected_data)

        CHAMADOS_DIR = os.path.join(os.path.dirname(__file__), '..', 'streamlit', 'chamados')
        os.makedirs(CHAMADOS_DIR, exist_ok=True)

        pdf_filename_for_email = "" # Usado para passar para o email
        excel_filename_for_email = os.path.join(CHAMADOS_DIR, "checklist_data.xlsx")

        attachments_for_email = [] # Lista de caminhos de arquivos para anexar

        success_messages = []
        error_messages = []

        # 1. Exportar para Excel
        try:
            self.export_to_excel(collected_data, excel_filename_for_email)
            success_messages.append(f"Dados salvos em '{excel_filename_for_email}'")
            if os.path.exists(excel_filename_for_email): # Adiciona à lista de anexos se o arquivo foi criado
                attachments_for_email.append(excel_filename_for_email)
        except Exception as e:
            print(f"Erro ao exportar para Excel: {e}")
            error_messages.append(f"Salvar em Excel: {e}")

        # 2. Gerar PDF
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            pdf_filename_for_email = os.path.join(CHAMADOS_DIR, f"checklist_{timestamp}.pdf")
            self.generate_pdf(collected_data, pdf_filename_for_email)
            success_messages.append(f"PDF gerado: '{pdf_filename_for_email}'")
            if os.path.exists(pdf_filename_for_email): # Adiciona à lista de anexos se o arquivo foi criado
                attachments_for_email.append(pdf_filename_for_email)
        except Exception as e:
            print(f"Erro ao gerar PDF: {e}")
            error_messages.append(f"Gerar PDF: {e}")
        
        # Adicionar a foto selecionada aos anexos do e-mail, se existir
        foto_path_incidente = self.incidentes_vars["foto_path"].get()
        if foto_path_incidente and os.path.exists(foto_path_incidente):
            attachments_for_email.append(foto_path_incidente)
            print(f"Foto '{foto_path_incidente}' adicionada aos anexos do e-mail.")
        elif foto_path_incidente: # Se há um caminho mas o arquivo não existe (ex: erro de cópia não tratado ou arquivo removido)
            print(f"Aviso: Arquivo da foto do incidente não encontrado em '{foto_path_incidente}', não será anexado.")
            # Opcional: Adicionar a error_messages para informar o usuário na UI
            # error_messages.append(f"Foto do incidente não encontrada em '{foto_path_incidente}'.")

        # A funcionalidade de e-mail foi removida/comentada.
        # # 3. Enviar E-mail (se houver anexos e um destinatário configurado)
        # # TODO: Adicionar um campo na UI para o email do destinatário ou ler de um arquivo de config.
        # # Por agora, usaremos um placeholder. Se estiver vazio, o email não será enviado.
        # recipient_email_placeholder = "seu_email_aqui@example.com" # ATENÇÃO: Configure isso!
        
        # if attachments_for_email and recipient_email_placeholder and "@" in recipient_email_placeholder:
        #     try:
        #         email_subject = f"Checklist de Inspeção - {collected_data['Dados da Visita']['Local da Subestação']} - {collected_data['Dados da Visita']['Data e Hora']}"
        #         email_body = "Segue em anexo o checklist de inspeção preenchido."
        #         # self.send_email_with_attachments( # CHAMADA COMENTADA
        #         #     recipient_email=recipient_email_placeholder,
        #         #     subject=email_subject,
        #         #     body=email_body,
        #         #     attachments=attachments_for_email
        #         # )
        #         # success_messages.append(f"E-mail enviado para '{recipient_email_placeholder}'") # MENSAGEM COMENTADA
        #     except Exception as e:
        #         print(f"Erro ao enviar e-mail: {e}")
        #         error_messages.append(f"Enviar E-mail: {e}")
        # elif not attachments_for_email and recipient_email_placeholder and "@" in recipient_email_placeholder :
        #      error_messages.append("E-mail não enviado: Nenhum anexo foi gerado com sucesso.")
        # elif not (recipient_email_placeholder and "@" in recipient_email_placeholder):
        #     print("Envio de e-mail pulado: Destinatário não configurado.")
        #     # Opcional: Adicionar uma mensagem a error_messages ou success_messages se quiser informar o usuário na UI
        #     # success_messages.append("Envio de e-mail pulado (destinatário não configurado).")

        # Exibir mensagens consolidadas
        if not error_messages and success_messages: # Se houve sucessos e nenhum erro
            messagebox.showinfo("Checklist Processado", "\n".join(success_messages))
        elif error_messages and success_messages: # Se houve sucessos parciais e erros
            full_error_msg = "Ocorreram erros:\n" + "\n".join(error_messages)
            messagebox.showwarning("Processamento Parcial", "\n".join(success_messages) + "\n\n" + full_error_msg)
        elif error_messages and not success_messages: # Se apenas erros ocorreram
            full_error_msg = "Ocorreram erros:\n" + "\n".join(error_messages)
            messagebox.showerror("Erro no Processamento", full_error_msg)
        elif not error_messages and not success_messages: # Caso raro, nada foi feito ou tentado
             messagebox.showinfo("Informação", "Nenhuma ação de exportação foi realizada.")


    def generate_pdf(self, data_dict, filename="checklist_output.pdf"):
        """Gera um PDF a partir dos dados do checklist."""
        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)
        
        # Título Principal
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, "Checklist de Inspeção da Subestação", 0, 1, "C")
        pdf.ln(5)

        for section_title, section_data in data_dict.items():
            pdf.set_font("Arial", "B", 12)
            pdf.cell(0, 10, section_title, 0, 1, "L")
            
            pdf.set_font("Arial", "", 10)
            if isinstance(section_data, dict):
                for key, value in section_data.items():
                    pdf.set_fill_color(240, 240, 240) # Cor de fundo leve para chaves
                    pdf.cell(60, 7, f"  {key}:", 0, 0, "L", fill= (section_title != "Observações Livres" and section_title != "Incidentes" and key != "Descrição") ) # Adiciona preenchimento para chaves
                    
                    # MultiCell para valores, especialmente para a descrição de incidentes
                    if section_title == "Incidentes" and key == "Descrição":
                        pdf.multi_cell(0, 7, str(value) if value else "-", 0, 1, "L")
                    else:
                         pdf.cell(0, 7, str(value) if value else "-", 0, 1, "L")
            else: # Para Observações Livres
                pdf.multi_cell(0, 7, str(section_data) if section_data else "-", 0, 1, "L")
            pdf.ln(3) # Espaço entre seções/itens

        try:
            pdf.output(filename, "F")
            print(f"PDF gerado com sucesso: '{filename}'")
        except Exception as e:
            raise Exception(f"Erro ao salvar PDF '{filename}': {e}")

    # def send_email_with_attachments(self, recipient_email, subject, body, attachments=None):
    #     """Envia um e-mail com os anexos especificados."""
    #     if attachments is None:
    #         attachments = []

    #     # ATENÇÃO: Substitua pelos seus dados reais de SMTP ou use um servidor de debug local.
    #     smtp_server = "localhost"  # Ou "smtp.example.com"
    #     smtp_port = 1025            # Porta para DebuggingServer. Para TLS: 587, SSL: 465
    #     smtp_user = ""              # Seu email, se o servidor exigir autenticação
    #     smtp_password = ""          # Sua senha, se o servidor exigir autenticação
    #     sender_email = "noreply_checklist@example.com" # Email remetente

    #     msg = MIMEMultipart()
    #     msg["From"] = sender_email
    #     msg["To"] = recipient_email
    #     msg["Subject"] = subject

    #     msg.attach(MIMEText(body, "plain"))

    #     for filepath in attachments:
    #         if not os.path.exists(filepath):
    #             print(f"Anexo não encontrado, pulando: {filepath}")
    #             continue
    #         try:
    #             with open(filepath, "rb") as attachment_file:
    #                 part = MIMEBase("application", "octet-stream")
    #                 part.set_payload(attachment_file.read())
    #             encoders.encode_base64(part)
    #             filename = os.path.basename(filepath)
    #             part.add_header(
    #                 "Content-Disposition",
    #                 f"attachment; filename= {filename}",
    #             )
    #             msg.attach(part)
    #             print(f"Anexado: {filename}")
    #         except Exception as e:
    #             print(f"Erro ao anexar o arquivo {filepath}: {e}")
    #             # Decide se quer levantar uma exceção ou apenas logar e continuar
    #             # raise Exception(f"Erro ao processar anexo {filepath}: {e}")

    #     try:
    #         server = smtplib.SMTP(smtp_server, smtp_port)
    #         # server.set_debuglevel(1) # Descomente para debug verboso do SMTP
    #         if smtp_user and smtp_password: # Se usar um servidor que não seja o de debug
    #              # server.starttls() # Descomente se o servidor usar TLS
    #              # server.login(smtp_user, smtp_password) # Descomente para fazer login
    #              pass # No servidor de debug, não precisa de TLS ou login
    #         server.sendmail(sender_email, recipient_email, msg.as_string())
    #         print(f"E-mail enviado com sucesso para {recipient_email}")
    #     except Exception as e:
    #         raise Exception(f"Falha ao enviar e-mail para {recipient_email}: {e}")
    #     finally:
    #         if 'server' in locals() and server:
    #             server.quit()


    def _flatten_data(self, data_dict):
        """Transforma o dicionário aninhado em um dicionário plano para o DataFrame."""
        flat_data = {}
        for section, items in data_dict.items():
            if isinstance(items, dict):
                for key, value in items.items():
                    flat_data[f"{section} - {key}"] = value
            else:
                flat_data[section] = items
        return flat_data

    def export_to_excel(self, data_dict, filename="checklist_data.xlsx"):
        """Exporta os dados do checklist para uma planilha Excel."""
        
        flat_data = self._flatten_data(data_dict)
        df_new_row = pd.DataFrame([flat_data]) # Cria DataFrame com uma linha

        try:
            # Tenta ler o arquivo existente
            df_existing = pd.read_excel(filename)
            # Concatena os dados existentes com a nova linha
            df_to_save = pd.concat([df_existing, df_new_row], ignore_index=True)
        except FileNotFoundError:
            # Se o arquivo não existe, o novo DataFrame é o que será salvo
            df_to_save = df_new_row
        except Exception as e:
            print(f"Erro ao ler o arquivo Excel existente '{filename}': {e}")
            # Se houver outro erro ao ler (ex: arquivo corrompido), decide-se salvar como novo
            # ou levantar o erro. Por segurança, vamos tentar salvar como novo se a leitura falhar.
            # Alternativamente, poderia levantar o erro: raise e
            messagebox.showwarning("Aviso ao Ler Excel", f"Não foi possível ler '{filename}'.\nUm novo arquivo será criado ou o existente pode ser sobrescrito com apenas os novos dados se a escrita falhar de forma diferente.\nDetalhe: {e}")
            df_to_save = df_new_row


        try:
            # Salva o DataFrame (novo ou concatenado) no arquivo Excel
            df_to_save.to_excel(filename, index=False)
            print(f"Dados salvos com sucesso em '{filename}'")
        except Exception as e:
            # Levanta a exceção para ser tratada na função submit_checklist
            raise Exception(f"Erro ao salvar dados no arquivo Excel '{filename}': {e}")


    def _collect_data(self):
        """Coleta todos os dados do formulário em um dicionário estruturado."""
        data = {
            "Dados da Visita": {
                "Data e Hora": self.data_hora_var.get(),
                "Local da Subestação": self.local_subestacao_var.get(),
                "Nome do Técnico": self.nome_tecnico_var.get(),
                "Tipo de Inspeção": self.tipo_inspecao_var.get(),
            },
            "Estado dos Equipamentos": {},
            "Segurança": {},
            "Incidentes": { # Inicializa para garantir que as chaves existam
                "Descrição": self.incidentes_vars["descricao"].get("1.0", tk.END).strip(),
                "Foto Anexada": self.incidentes_vars["foto_path"].get() if self.incidentes_vars["foto_path"].get() else "Nenhuma",
                "Gravidade": self.incidentes_vars["gravidade"].get(),
                "Ação Tomada": self.incidentes_vars["acao_tomada"].get()
            },
            "Observações Livres": self.observacoes_text.get("1.0", tk.END).strip() or "Nenhuma observação."
        }

        for key, var in self.estado_equipamentos_vars.items():
            # Limpa o nome da chave para melhor leitura
            clean_key = key.replace('_transformador', '').replace('_disjuntor', '').replace('_barramento', '').replace('_', ' ').capitalize()
            data["Estado dos Equipamentos"][clean_key] = var.get()

        for key, var in self.seguranca_vars.items():
            clean_key = key.replace('_', ' ').capitalize()
            data["Segurança"][clean_key] = "Verificado" if var.get() else "Não Verificado"
        
        return data

if __name__ == "__main__":
    root = tk.Tk()
    app = ChecklistApp(root)
    root.mainloop()
